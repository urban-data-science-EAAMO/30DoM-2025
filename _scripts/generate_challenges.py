import re
from datetime import date, datetime
from pathlib import Path
from typing import List, Optional

import openpyxl


def is_date_like(value: object) -> bool:
    return isinstance(value, (datetime, date))


def try_format_date(value: object) -> Optional[str]:
    if isinstance(value, datetime):
        return value.date().strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        text = value.strip()
        # Accept common formats like MM-DD-YYYY or M-D-YYYY
        m = re.match(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$", text)
        if m:
            mm, dd, yyyy = m.groups()
            try:
                dt = datetime(int(yyyy), int(mm), int(dd))
                return dt.strftime("%Y-%m-%d")
            except ValueError:
                return None
        return None
    # Fall back to None for non-datetime; the README will still show raw text
    return None


def safe_filename_component(text: str, max_length: int = 80) -> str:
    if not text:
        return "unknown"
    # Replace path separators and control characters
    cleaned = re.sub(r"[\\/\0\n\r\t]+", "-", text)
    # Replace characters that are problematic on common filesystems
    cleaned = re.sub(r"[:*?\"<>|]+", "-", cleaned)
    # Collapse spaces
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    # Trim length
    if len(cleaned) > max_length:
        cleaned = cleaned[:max_length].rstrip()
    return cleaned


DATE_HEADER = "Date"
TITLE_HEADER = "Challenge Name"
DESCRIPTION_HEADER = "Description"


def get_headers(ws) -> List[str]:
    # Headers are on the second row per provided sheet format
    return [str(c.value).strip() if c.value is not None else "" for c in ws[2]]


def extract_members(values: List[object], member_cols: List[int]) -> List[str]:
    members: List[str] = []
    for col in member_cols:
        if col >= len(values):
            continue
        cell_val = values[col]
        if cell_val is None:
            continue
        text = str(cell_val).strip()
        if not text:
            continue
        # Split common separators
        parts = re.split(r"[,;\n]+", text)
        for p in parts:
            name = p.strip()
            if name and name.lower() not in {"n/a", "na", "none", "-"}:
                members.append(name)
    # De-duplicate while preserving order
    seen = set()
    deduped: List[str] = []
    for m in members:
        if m.lower() in seen:
            continue
        seen.add(m.lower())
        deduped.append(m)
    return deduped


def generate_readme_content(
    title: str,
    date_text: Optional[str],
    description: Optional[str],
    members: List[str],
) -> str:
    lines: List[str] = []
    lines.append(f"# {title}")
    if date_text:
        lines.append("")
        lines.append(f"**Date**: {date_text}")
    if description:
        lines.append("")
        lines.append("**Description**:")
        lines.append("")
        lines.append(description)
    if members:
        lines.append("")
        lines.append("**Members**:")
        for m in members:
            lines.append(f"- {m}")
    lines.append("")
    return "\n".join(lines)


def main() -> int:
    repo_root = Path(__file__).resolve().parent.parent
    signup_path = repo_root / "_signup_sheet" / "30 Days of Mapping Sign-up.xlsx"
    output_root = repo_root / "challenges"

    if not signup_path.exists():
        print(f"Signup sheet not found at: {signup_path}")
        return 1

    output_root.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(signup_path, data_only=True)
    ws = wb.active

    headers = get_headers(ws)
    normalized_headers = [h.strip().lower() for h in headers]
    try:
        date_col = normalized_headers.index(DATE_HEADER.lower())
    except ValueError:
        print(f"Missing required header: '{DATE_HEADER}' in row 1")
        return 1
    try:
        title_col = normalized_headers.index(TITLE_HEADER.lower())
    except ValueError:
        print(f"Missing required header: '{TITLE_HEADER}' in row 1")
        return 1
    description_col = (
        normalized_headers.index(DESCRIPTION_HEADER.lower())
        if DESCRIPTION_HEADER.lower() in normalized_headers
        else None
    )
    # All other columns (except the three above) are member columns
    member_cols = [
        i
        for i, _ in enumerate(headers)
        if i not in {date_col, title_col} and (description_col is None or i != description_col)
    ]

    created_count = 0
    updated_count = 0

    # Data starts from row 3 (row 1 = title, row 2 = headers)
    for r in range(3, ws.max_row + 1):
        row_cells = ws[r]
        values = [c.value for c in row_cells]

        # Skip completely empty rows
        if not any(v is not None and str(v).strip() for v in values):
            continue

        title_val = values[title_col] if title_col < len(values) else None
        title_text = str(title_val or "").strip()
        if not title_text:
            # No challenge title → skip row
            continue

        date_val = values[date_col] if date_col is not None and date_col < len(values) else None
        description_val = (
            values[description_col]
            if description_col is not None and description_col < len(values)
            else None
        )
        description_text = str(description_val).strip() if description_val is not None else None

        members = extract_members(values, member_cols)

        # Folder name: optional date prefix + sanitized title
        date_for_folder = try_format_date(date_val)
        title_component = safe_filename_component(title_text, max_length=80)
        if date_for_folder:
            folder_name = f"{date_for_folder} - {title_component}"
        else:
            folder_name = title_component

        folder_path = output_root / folder_name
        folder_path.mkdir(parents=True, exist_ok=True)

        # README content
        date_text = (
            try_format_date(date_val)
            if is_date_like(date_val)
            else (str(date_val).strip() if date_val is not None else None)
        )
        readme_content = generate_readme_content(
            title=title_text,
            date_text=date_text,
            description=description_text,
            members=members,
        )

        readme_path = folder_path / "README.md"
        existing = readme_path.read_text(encoding="utf-8") if readme_path.exists() else None
        if existing != readme_content:
            readme_path.write_text(readme_content, encoding="utf-8")
            if existing is None:
                created_count += 1
            else:
                updated_count += 1

    print(
        f"Generation complete. Created {created_count} new README(s), updated {updated_count} README(s)."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


